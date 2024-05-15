import ifcopenshell
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import re
from tqdm import trange
from time import sleep

def get_entities_filtered(ifcschema_entities, get_types):
    ents_not_collect = ["IfcGeometricRepresentationItem", "Ifcobject", "IfcobjectDefinition", "IfcProduct", "IfcRelationship", "IfcRepresentationItem", "IfcRoot"]
    if get_types:
        ents = [e for e in ifcschema_entities if "type" in e.lower()]
    else:
        ents = [e for e in ifcschema_entities if "type" not in e.lower() and not e in ents_not_collect]
    return (len(ents), ents)

def get_ents_info_to_df(ifc_entName):
    # get all given entities info
    entity_info = [en.get_info() for en in ifc_file.by_type(ifc_entName)]
    #print(entity_info)
    #convert data to a pandas dataframe
    df = pd.DataFrame(entity_info)
    return df

def contract_entName(entName, trunc_n = 3):
    # takes a Ifc entity name, split it at each uppercase occurrence, then slice each portion taking its trunc_n leftmost chars, finally concatenates back the contracted Ifc entity name in a string
    entName_split = re.findall('[A-Z][^A-Z]*', entName)
    return "".join([s[:trunc_n] if len(s) >= trunc_n else s for s in entName_split])

def create_ws(wb, ws_name):
    # Check if the sheet exists
    if ws_name in wb.sheetnames:
        ws = wb.create_sheet(title=ws_name + "_1")
    else:
        ws = wb.create_sheet(title=ws_name)
    return ws

def create_ws_and_table(wb, ifc_entName):
    # Get entity info DataFrame
    df = get_ents_info_to_df(ifc_entName)
    
    # Reduce Lengthy entity names in contract version
    if len(ifc_entName) > 30: ifc_entName = contract_entName(ifc_entName, trunc_n = 3)

    # Create a new worksheet named as per Ifc_entName
    ws = create_ws(wb, ifc_entName)

    # Write the DataFrame to the Excel file
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append([str(v) if str(v).startswith('#') or '(' in str(v) else v for v in row])

    # Get dataframe shape
    df_shape = df.shape

    # Set table Excel range
    tbl_xl_rangeAddress = "A1:" + cols_dict[df_shape[1]-1] + str(df_shape[0]+1)

    # Create an Excel table
    tbl = Table(displayName=ifc_entName, ref=tbl_xl_rangeAddress)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tbl.tableStyleInfo = style

    # Commit table to worksheet
    ws.add_table(tbl)

    return True

def remove_ws(wb, ws_name="sheet"):
    # Check if the sheet exists
    if ws_name in wb.sheetnames:
        # Get the sheet to delete
        ws_to_delete = wb[ws_name]
        # Delete the sheet
        wb.remove(ws_to_delete)
        print(f"Sheet {ws_name} has been deleted.")

def purge_wb(wb):
    # Iterate through each sheet
    for ws in wb.sheetnames:
        ws = wb[ws]
        # Remove all tables
        for tbl in ws.tables.values():
            ws._tables.remove(tbl)
        # Remove wS
        remove_ws(wb, ws_name=ws)
    return True

def generate_column_names():
    from itertools import product
    import string
    cols = list(string.ascii_uppercase)
    for p in product(string.ascii_uppercase, repeat=2):
        cols.append(''.join(p))
    return cols

# INPUT

import tkinter as tk
from tkinter import filedialog

def select_ifc_file():
    # Create a Tkinter root window and hide it
    root = tk.Tk()
    root.withdraw()

    # Open the file dialog to select an IFC file
    file_path = filedialog.askopenfilename(
        title="Select IFC file",
        filetypes=[("IFC files", "*.ifc")]
    )

    # Split the file path into base path and file name
    if file_path:
        ifc_basepath = r"{}".format(file_path.rsplit('/', 1)[0]).replace('/', '\\')
        ifc_filename = file_path.rsplit('/', 1)[1].rsplit('.', 1)[0]
        return ifc_basepath, ifc_filename
    else:
        return None, None

# Run the file selection function
ifc_basepath, ifc_filename = select_ifc_file()

# Print the results for verification
ifc_getTypes = False

# CODE

# Load model
ifc_file = ifcopenshell.open(f"{ifc_basepath}\\{ifc_filename}.ifc")

sch_entities_names = [e.name() for e in ifcopenshell.schema_by_name(ifc_file.schema).entities()]

# Add error handling for missing entities
entities_names_in_use = []
for en in sch_entities_names:
    try:
        if len(ifc_file.by_type(en)) != 0:
            entities_names_in_use.append(en)
    except RuntimeError as e:
        print(f"Entity {en} not found in schema.")
# Filtered entities names
ents = get_entities_filtered(entities_names_in_use, get_types=ifc_getTypes)
ents

# Construct output Excel filename

# Define the base directory. This can be set explicitly or derived from the script's directory.
base_dir = ifc_basepath

# Construct the path to the Excel file using the base directory
xls_filename = os.path.join(
    base_dir,
    f"{datetime.now().strftime('%Y-%m-%d')}_{ifc_filename}_entTypes_{ifc_getTypes}_2.xlsx"
)

# Create dictionary of alphabetical letters (to be used to set table Excel range)
cols_dict = dict(enumerate(generate_column_names()))
cols_dict

# Create a new Excel workbook
wb = Workbook()

# Remove all sheets and tables
purge_wb(wb)

# Save the Excel file
wb.save(xls_filename)

# Add a new worksheet and a table with all entities data, showing a progress bar
t = trange(len(ents[1]), desc='Entity:', leave=True)
for i, en in zip(t, ents[1]):
    t.set_description(f"Entity: {en}")
    t.refresh()  # to show immediately the update
    create_ws_and_table(wb, en)

# Remove default worksheet named "sheet" if present
remove_ws(wb, ws_name="sheet")

# Save the Excel file
wb.save(xls_filename)
print(f"File {xls_filename} has been saved successfully!")